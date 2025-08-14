from colegios.models import Colegio
from cursos.models import Curso
from estudiantes.models import Estudiante

from atrasos.models import Atraso
from salidas.models import Salida
import matplotlib.pyplot as plt
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models.functions import TruncDate, TruncMonth, TruncYear, TruncWeek
from django.template.loader import render_to_string
import json
import hashlib
import csv
from django.db import connection
from collections import defaultdict

# Importaciones para exportación
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference

# ReportLab para PDF (más compatible con Windows)
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import Image as RLImage

# Matplotlib para gráficos en PDF (backend sin UI)
import matplotlib
matplotlib.use('Agg')


@login_required
def dashboard_analytics(request):
    """Dashboard principal de analytics con filtros dinámicos"""

    # Obtener filtros de la URL
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    curso_id = request.GET.get('curso')
    tipo_incidencia = request.GET.get('tipo_incidencia')

    # Aplicar filtros base para atrasos y salidas
    filtros_atrasos = Q()
    filtros_salidas = Q()

    if fecha_inicio:
        filtros_atrasos &= Q(fecha__gte=fecha_inicio)
        filtros_salidas &= Q(fecha__gte=fecha_inicio)

    if fecha_fin:
        filtros_atrasos &= Q(fecha__lte=fecha_fin)
        filtros_salidas &= Q(fecha__lte=fecha_fin)

    if curso_id:
        filtros_atrasos &= Q(estudiante__curso_id=curso_id)
        filtros_salidas &= Q(estudiante__curso_id=curso_id)

    # Aplicar filtro de tipo de incidencia
    if tipo_incidencia == 'atrasos':
        filtros_salidas = Q(pk__isnull=True)  # No mostrar salidas
    elif tipo_incidencia == 'salidas':
        filtros_atrasos = Q(pk__isnull=True)  # No mostrar atrasos

    # Obtener datos de atrasos y salidas
    atrasos = Atraso.objects.filter(filtros_atrasos)
    salidas = Salida.objects.filter(filtros_salidas)

    # Métricas principales
    total_atrasos = atrasos.count()
    atrasos_justificados = atrasos.filter(justificado=True).count()
    atrasos_no_justificados = atrasos.filter(justificado=False).count()
    total_salidas = salidas.count()

    # Porcentajes
    porcentaje_justificados = (
        atrasos_justificados / total_atrasos * 100) if total_atrasos > 0 else 0
    porcentaje_no_justificados = (
        atrasos_no_justificados / total_atrasos * 100) if total_atrasos > 0 else 0

    # Análisis por curso
    atrasos_por_curso = atrasos.values('estudiante__curso__nombre').annotate(
        total=Count('id'),
        justificados=Count('id', filter=Q(justificado=True)),
        no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total')

    salidas_por_curso = salidas.values('estudiante__curso__nombre').annotate(
        total=Count('id')
    ).order_by('-total')

    # Análisis temporal
    # Agrupamiento temporal compatible con SQLite y PostgreSQL
    if connection.vendor == 'sqlite':
        # Agrupar atrasos por fecha en Python
        atrasos_por_fecha_dict = defaultdict(
            lambda: {'total': 0, 'justificados': 0, 'no_justificados': 0})
        for atraso in atrasos:
            fecha = atraso.fecha
            atrasos_por_fecha_dict[fecha]['total'] += 1
            if atraso.justificado:
                atrasos_por_fecha_dict[fecha]['justificados'] += 1
            else:
                atrasos_por_fecha_dict[fecha]['no_justificados'] += 1
        atrasos_por_fecha = [
            {'fecha_simple': fecha, **data}
            for fecha, data in sorted(atrasos_por_fecha_dict.items())
        ]

    else:
        atrasos_por_fecha = atrasos.annotate(
            fecha_simple=TruncDate('fecha')
        ).values('fecha_simple').annotate(
            total=Count('id'),
            justificados=Count('id', filter=Q(justificado=True)),
            no_justificados=Count('id', filter=Q(justificado=False))
        ).order_by('fecha_simple')

    # Análisis temporal de salidas
    if connection.vendor == 'sqlite':
        # Agrupar salidas por fecha en Python
        salidas_por_fecha_dict = defaultdict(lambda: {'total': 0})
        for salida in salidas:
            fecha = salida.fecha
            salidas_por_fecha_dict[fecha]['total'] += 1
        salidas_por_fecha = [
            {'fecha_simple': fecha, **data}
            for fecha, data in sorted(salidas_por_fecha_dict.items())
        ]
    else:
        salidas_por_fecha = salidas.annotate(
            fecha_simple=TruncDate('fecha')
        ).values('fecha_simple').annotate(
            total=Count('id')
        ).order_by('fecha_simple')

    # Top estudiantes con más incidencias
    top_estudiantes_atrasos = atrasos.values(
        'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre'
    ).annotate(
        total_atrasos=Count('id'),
        atrasos_justificados=Count('id', filter=Q(justificado=True)),
        atrasos_no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total_atrasos')[:10]

    top_estudiantes_salidas = salidas.values(
        'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre'
    ).annotate(
        total_salidas=Count('id')
    ).order_by('-total_salidas')[:10]

    # Análisis por hora del día
    atrasos_por_hora = atrasos.values('hora').annotate(
        total=Count('id')
    ).order_by('hora')

    # Preparar datos para JavaScript - convertir fechas a strings y manejar valores nulos
    def prepare_data_for_js(data_list):
        """Convierte los datos para que sean compatibles con JavaScript"""
        prepared_data = []
        for item in data_list:
            prepared_item = {}
            for key, value in item.items():
                if hasattr(value, 'strftime'):  # Es una fecha
                    prepared_item[key] = value.strftime('%Y-%m-%d')
                elif value is None:
                    prepared_item[key] = ''
                else:
                    prepared_item[key] = value
            prepared_data.append(prepared_item)
        return prepared_data

    # Contexto para el template
    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'curso_id': curso_id,
        'tipo_incidencia': tipo_incidencia,

        # Métricas principales
        'total_atrasos': total_atrasos,
        'atrasos_justificados': atrasos_justificados,
        'atrasos_no_justificados': atrasos_no_justificados,
        'porcentaje_justificados': round(porcentaje_justificados, 1),
        'porcentaje_no_justificados': round(porcentaje_no_justificados, 1),
        'total_salidas': total_salidas,

        # Datos para gráficos - serializados para JavaScript
        'atrasos_por_curso': json.dumps(prepare_data_for_js(list(atrasos_por_curso))),
        'atrasos_por_fecha': json.dumps(prepare_data_for_js(list(atrasos_por_fecha))),
        'atrasos_por_hora': json.dumps(prepare_data_for_js(list(atrasos_por_hora))),
        'salidas_por_curso': json.dumps(prepare_data_for_js(list(salidas_por_curso))),
        'salidas_por_fecha': json.dumps(prepare_data_for_js(list(salidas_por_fecha))),

        # Top estudiantes
        'top_estudiantes_atrasos': list(top_estudiantes_atrasos),
        'top_estudiantes_salidas': list(top_estudiantes_salidas),

        # Filtros disponibles
        'cursos': Curso.objects.all().order_by('nombre'),
    }

    return render(request, 'reportes/dashboard_analytics.html', context)


@login_required
def api_datos_graficos(request):
    """API para obtener datos de gráficos en formato JSON"""

    # Obtener parámetros
    tipo_grafico = request.GET.get('tipo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    curso_id = request.GET.get('curso')

    # Aplicar filtros solo para atrasos
    filtros_atrasos = Q()

    if fecha_inicio:
        filtros_atrasos &= Q(fecha__gte=fecha_inicio)

    if fecha_fin:
        filtros_atrasos &= Q(fecha__lte=fecha_fin)

    if curso_id:
        filtros_atrasos &= Q(estudiante__curso_id=curso_id)

    atrasos = Atraso.objects.filter(filtros_atrasos)

    if tipo_grafico == 'atrasos_por_curso':
        datos = atrasos.values('estudiante__curso__nombre').annotate(
            total=Count('id'),
            justificados=Count('id', filter=Q(justificado=True)),
            no_justificados=Count('id', filter=Q(justificado=False))
        ).order_by('-total')

    elif tipo_grafico == 'atrasos_temporales':
        datos = atrasos.annotate(
            fecha_simple=TruncDate('fecha')
        ).values('fecha_simple').annotate(
            total=Count('id'),
            justificados=Count('id', filter=Q(justificado=True)),
            no_justificados=Count('id', filter=Q(justificado=False))
        ).order_by('fecha_simple')

    elif tipo_grafico == 'atrasos_por_hora':
        datos = atrasos.values('hora').annotate(
            total=Count('id')
        ).order_by('hora')

    else:
        datos = []

    return JsonResponse({'datos': list(datos)})


@login_required
def reporte_detallado(request):
    """Reporte detallado con análisis profundo"""

    # Obtener filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    curso_id = request.GET.get('curso')
    tipo_analisis = request.GET.get('tipo_analisis', 'general')

    # Aplicar filtros solo para atrasos
    filtros_atrasos = Q()

    if fecha_inicio:
        filtros_atrasos &= Q(fecha__gte=fecha_inicio)

    if fecha_fin:
        filtros_atrasos &= Q(fecha__lte=fecha_fin)

    if curso_id:
        filtros_atrasos &= Q(estudiante__curso_id=curso_id)

    atrasos = Atraso.objects.filter(filtros_atrasos)

    # Análisis detallado según tipo
    if tipo_analisis == 'tendencias':
        if connection.vendor == 'sqlite':
            # Agrupar atrasos por mes en Python
            atrasos_tendencias_dict = defaultdict(
                lambda: {'total': 0, 'justificados': 0, 'no_justificados': 0})
            for atraso in atrasos:
                mes = atraso.fecha.replace(day=1)
                atrasos_tendencias_dict[mes]['total'] += 1
                if atraso.justificado:
                    atrasos_tendencias_dict[mes]['justificados'] += 1
                else:
                    atrasos_tendencias_dict[mes]['no_justificados'] += 1
            atrasos_tendencias = [
                {'mes': mes, **data}
                for mes, data in sorted(atrasos_tendencias_dict.items())
            ]

        else:
            atrasos_tendencias = atrasos.annotate(
                mes=TruncMonth('fecha')
            ).values('mes').annotate(
                total=Count('id'),
                justificados=Count('id', filter=Q(justificado=True)),
                no_justificados=Count('id', filter=Q(justificado=False))
            ).order_by('mes')

        datos_analisis = {
            'atrasos_tendencias': list(atrasos_tendencias),
        }

    elif tipo_analisis == 'comportamiento':
        # Análisis de comportamiento por estudiante
        comportamiento_estudiantes = atrasos.values(
            'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre'
        ).annotate(
            total_atrasos=Count('id'),
            atrasos_justificados=Count('id', filter=Q(justificado=True)),
            atrasos_no_justificados=Count('id', filter=Q(justificado=False)),
            porcentaje_justificados=Avg('justificado') * 100
        ).order_by('-total_atrasos')

        datos_analisis = {
            'comportamiento_estudiantes': list(comportamiento_estudiantes),
        }

    else:  # general
        # Análisis general
        datos_analisis = {
            'total_atrasos': atrasos.count(),
            'atrasos_justificados': atrasos.filter(justificado=True).count(),
            'atrasos_no_justificados': atrasos.filter(justificado=False).count(),
        }

    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'curso_id': curso_id,
        'tipo_analisis': tipo_analisis,
        'datos_analisis': datos_analisis,
        'cursos': Curso.objects.all().order_by('nombre'),
    }

    return render(request, 'reportes/reporte_detallado.html', context)


# Funciones de exportación


@login_required
def exportar_reporte_pdf(request):
    """Exportar reporte en formato PDF"""
    # Obtener los mismos datos que el dashboard
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    curso_id = request.GET.get('curso')
    tipo_incidencia = request.GET.get('tipo_incidencia', 'todos')

    # Aplicar filtros base solo para atrasos
    filtros_atrasos = Q()

    if fecha_inicio:
        filtros_atrasos &= Q(fecha__gte=fecha_inicio)

    if fecha_fin:
        filtros_atrasos &= Q(fecha__lte=fecha_fin)

    if curso_id:
        filtros_atrasos &= Q(estudiante__curso_id=curso_id)

    # Obtener datos
    atrasos = Atraso.objects.filter(filtros_atrasos)

    # Métricas principales
    total_atrasos = atrasos.count()
    atrasos_justificados = atrasos.filter(justificado=True).count()
    atrasos_no_justificados = atrasos.filter(justificado=False).count()

    # Porcentajes
    porcentaje_justificados = (
        atrasos_justificados / total_atrasos * 100) if total_atrasos > 0 else 0
    porcentaje_no_justificados = (
        atrasos_no_justificados / total_atrasos * 100) if total_atrasos > 0 else 0

    # Análisis por curso
    atrasos_por_curso = atrasos.values('estudiante__curso__nombre').annotate(
        total=Count('id'),
        justificados=Count('id', filter=Q(justificado=True)),
        no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total')

    # Top estudiantes
    top_estudiantes_atrasos = atrasos.values(
        'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre'
    ).annotate(
        total_atrasos=Count('id'),
        atrasos_justificados=Count('id', filter=Q(justificado=True)),
        atrasos_no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total_atrasos')[:10]

    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_atrasos': total_atrasos,
        'atrasos_justificados': atrasos_justificados,
        'atrasos_no_justificados': atrasos_no_justificados,
        'porcentaje_justificados': round(porcentaje_justificados, 1),
        'porcentaje_no_justificados': round(porcentaje_no_justificados, 1),
        'atrasos_por_curso': list(atrasos_por_curso),
        'top_estudiantes_atrasos': list(top_estudiantes_atrasos),
        'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M'),
    }

    # Generar PDF con ReportLab
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#366092')
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor('#366092')
    )

    # Título principal
    story.append(
        Paragraph("REPORTE DE ANALYTICS - ATRASOS", title_style))
    story.append(Spacer(1, 20))

    # Filtros aplicados
    story.append(Paragraph("Filtros Aplicados:", heading_style))
    curso_nombre = 'Todos los cursos'
    if curso_id:
        try:
            curso_obj = Curso.objects.filter(id=curso_id).first()
            if curso_obj:
                curso_nombre = curso_obj.nombre
            else:
                curso_nombre = f"ID {curso_id}"
        except Exception:
            curso_nombre = f"ID {curso_id}"
    filtros_data = [
        ['Fecha inicio:', fecha_inicio or 'Todas las fechas'],
        ['Fecha fin:', fecha_fin or 'Todas las fechas'],
        ['Curso:', curso_nombre],
        ['Fecha generación:', timezone.now().strftime('%d/%m/%Y %H:%M')]
    ]
    filtros_table = Table(filtros_data, colWidths=[2*inch, 4*inch])
    filtros_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(filtros_table)
    story.append(Spacer(1, 20))

    # Métricas principales en tabla compacta
    story.append(Spacer(1, 30))
    story.append(Paragraph("Métricas Principales", heading_style))
    story.append(Spacer(1, 20))

    # Tabla de métricas principales
    metrics_data = [
        ['Métrica', 'Valor'],
        ['Total Atrasos', str(total_atrasos)],
        ['Atrasos Justificados', f"{round(porcentaje_justificados, 1)}%"],
        ['Atrasos No Justificados', f"{round(porcentaje_no_justificados, 1)}%"]
    ]

    metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        # Solo el encabezado en blanco
        ('TEXTCOLOR', (0, 0), (0, 0), colors.whitesmoke),
        # El resto del contenido en negro
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#f9f9f9')]),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 30))

    # Gráficos principales (embebidos como imágenes)

    def build_chart_atrasos_por_curso(data):
        labels = [item.get('estudiante__curso__nombre')
                  or 'Curso desconocido' for item in data][:10]
        justificados = [int(item.get('justificados') or 0)
                        for item in data][:10]
        no_justificados = [int(item.get('no_justificados') or 0)
                           for item in data][:10]

        if not labels:
            return None

        fig, ax = plt.subplots(figsize=(8.5, 4))
        ax.bar(labels, justificados, label='Justificados', color='#28a745')
        ax.bar(labels, no_justificados, bottom=justificados,
               label='No Justificados', color='#dc3545')
        ax.set_title('Atrasos por Curso (Top 10)')
        ax.set_ylabel('Cantidad')
        ax.set_xlabel('Curso')
        ax.legend(loc='upper right')
        ax.tick_params(axis='x', rotation=35, labelsize=8)
        ax.grid(axis='y', linestyle='--', alpha=0.3)
        plt.tight_layout()

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig)
        img_buffer.seek(0)
        return img_buffer

    chart1 = build_chart_atrasos_por_curso(list(atrasos_por_curso))
    story.append(Spacer(1, 200))

    story.append(Paragraph("Visualizaciones", heading_style))
    if chart1:
        story.append(Paragraph("Atrasos por Curso", styles['Normal']))
        story.append(Spacer(1, 6))
        story.append(RLImage(chart1, width=6.3*inch, height=3.6*inch))
        story.append(Spacer(1, 16))

    # Atrasos por curso
    story.append(Spacer(1, 30))
    story.append(Paragraph("Atrasos por Curso", heading_style))
    if atrasos_por_curso:
        atrasos_curso_data = [
            ['Curso', 'Total', 'Justificados', 'No Justificados']]
        for item in atrasos_por_curso:
            atrasos_curso_data.append([
                item['estudiante__curso__nombre'] or 'Curso desconocido',
                str(item['total']),
                str(item['justificados']),
                str(item['no_justificados'])
            ])
        atrasos_curso_table = Table(atrasos_curso_data, colWidths=[
                                    2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        atrasos_curso_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        story.append(atrasos_curso_table)
    else:
        story.append(Paragraph("No hay datos disponibles", styles['Normal']))
    story.append(Spacer(1, 20))

    story.append(PageBreak())

    # === Estadísticas por periodo (Semana / Mes / Año) ===
    def compute_periodic_rankings(qs, group_field_label):
        """Devuelve dict con rankings por semana/mes/año para estudiantes y cursos.
        group_field_label: 'estudiante' o 'curso' para etiquetar cabeceras.
        """
        results = {}
        if connection.vendor == 'sqlite':
            # Agrupación en Python para compatibilidad
            def week_start(d):
                return d - timedelta(days=d.weekday())

            # Semanal
            weekly_counter_students = defaultdict(lambda: defaultdict(
                lambda: {'total': 0, 'justificados': 0, 'no_justificados': 0}))
            weekly_counter_courses = defaultdict(
                lambda: defaultdict(lambda: {'total': 0}))
            for a in qs:
                w = week_start(a.fecha)
                student_key = (a.estudiante.nombre, a.estudiante.rut, getattr(
                    a.estudiante.curso, 'nombre', 'Curso desconocido'))
                course_key = getattr(a.estudiante.curso,
                                     'nombre', 'Curso desconocido')
                weekly_counter_students[w][student_key]['total'] += 1
                if a.justificado:
                    weekly_counter_students[w][student_key]['justificados'] += 1
                else:
                    weekly_counter_students[w][student_key]['no_justificados'] += 1
                weekly_counter_courses[w][course_key]['total'] += 1

            weekly_students = []
            for period, students in weekly_counter_students.items():
                for (nombre, rut, curso), c in students.items():
                    weekly_students.append({'periodo': period, 'estudiante__nombre': nombre, 'estudiante__rut': rut, 'estudiante__curso__nombre': curso,
                                           'total': c['total'], 'justificados': c['justificados'], 'no_justificados': c['no_justificados']})
            weekly_students.sort(key=lambda x: x['total'], reverse=True)
            results['weekly_students'] = weekly_students[:10]

            weekly_courses = []
            for period, courses in weekly_counter_courses.items():
                for curso, c in courses.items():
                    weekly_courses.append(
                        {'periodo': period, 'curso__nombre': curso, 'total': c['total']})
            weekly_courses.sort(key=lambda x: x['total'], reverse=True)
            results['weekly_courses'] = weekly_courses[:10]

            # Mensual
            monthly_counter_students = defaultdict(lambda: defaultdict(
                lambda: {'total': 0, 'justificados': 0, 'no_justificados': 0}))
            monthly_counter_courses = defaultdict(
                lambda: defaultdict(lambda: {'total': 0}))
            for a in qs:
                m = a.fecha.replace(day=1)
                student_key = (a.estudiante.nombre, a.estudiante.rut, getattr(
                    a.estudiante.curso, 'nombre', 'Curso desconocido'))
                course_key = getattr(a.estudiante.curso,
                                     'nombre', 'Curso desconocido')
                monthly_counter_students[m][student_key]['total'] += 1
                if a.justificado:
                    monthly_counter_students[m][student_key]['justificados'] += 1
                else:
                    monthly_counter_students[m][student_key]['no_justificados'] += 1
                monthly_counter_courses[m][course_key]['total'] += 1

            monthly_students = []
            for period, students in monthly_counter_students.items():
                for (nombre, rut, curso), c in students.items():
                    monthly_students.append({'periodo': period, 'estudiante__nombre': nombre, 'estudiante__rut': rut, 'estudiante__curso__nombre': curso,
                                            'total': c['total'], 'justificados': c['justificados'], 'no_justificados': c['no_justificados']})
            monthly_students.sort(key=lambda x: x['total'], reverse=True)
            results['monthly_students'] = monthly_students[:10]

            monthly_courses = []
            for period, courses in monthly_counter_courses.items():
                for curso, c in courses.items():
                    monthly_courses.append(
                        {'periodo': period, 'curso__nombre': curso, 'total': c['total']})
            monthly_courses.sort(key=lambda x: x['total'], reverse=True)
            results['monthly_courses'] = monthly_courses[:10]

            # Anual
            yearly_counter_students = defaultdict(lambda: defaultdict(
                lambda: {'total': 0, 'justificados': 0, 'no_justificados': 0}))
            yearly_counter_courses = defaultdict(
                lambda: defaultdict(lambda: {'total': 0}))
            for a in qs:
                y = a.fecha.replace(month=1, day=1)
                student_key = (a.estudiante.nombre, a.estudiante.rut, getattr(
                    a.estudiante.curso, 'nombre', 'Curso desconocido'))
                course_key = getattr(a.estudiante.curso,
                                     'nombre', 'Curso desconocido')
                yearly_counter_students[y][student_key]['total'] += 1
                if a.justificado:
                    yearly_counter_students[y][student_key]['justificados'] += 1
                else:
                    yearly_counter_students[y][student_key]['no_justificados'] += 1
                yearly_counter_courses[y][course_key]['total'] += 1

            yearly_students = []
            for period, students in yearly_counter_students.items():
                for (nombre, rut, curso), c in students.items():
                    yearly_students.append({'periodo': period, 'estudiante__nombre': nombre, 'estudiante__rut': rut, 'estudiante__curso__nombre': curso,
                                           'total': c['total'], 'justificados': c['justificados'], 'no_justificados': c['no_justificados']})
            yearly_students.sort(key=lambda x: x['total'], reverse=True)
            results['yearly_students'] = yearly_students[:10]

            yearly_courses = []
            for period, courses in yearly_counter_courses.items():
                for curso, c in courses.items():
                    yearly_courses.append(
                        {'periodo': period, 'curso__nombre': curso, 'total': c['total']})
            yearly_courses.sort(key=lambda x: x['total'], reverse=True)
            results['yearly_courses'] = yearly_courses[:10]

        else:
            # Backend con funciones de truncado
            results['weekly_students'] = list(qs.annotate(
                periodo=TruncWeek('fecha')
            ).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'),
                justificados=Count('id', filter=Q(justificado=True)),
                no_justificados=Count('id', filter=Q(justificado=False))
            ).order_by('-total')[:10])

            results['monthly_students'] = list(qs.annotate(
                periodo=TruncMonth('fecha')
            ).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'),
                justificados=Count('id', filter=Q(justificado=True)),
                no_justificados=Count('id', filter=Q(justificado=False))
            ).order_by('-total')[:10])

            results['yearly_students'] = list(qs.annotate(
                periodo=TruncYear('fecha')
            ).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'),
                justificados=Count('id', filter=Q(justificado=True)),
                no_justificados=Count('id', filter=Q(justificado=False))
            ).order_by('-total')[:10])

            results['weekly_courses'] = list(qs.annotate(
                periodo=TruncWeek('fecha')
            ).values('periodo', 'estudiante__curso__nombre').annotate(
                total=Count('id')
            ).order_by('-total')[:10])

            results['monthly_courses'] = list(qs.annotate(
                periodo=TruncMonth('fecha')
            ).values('periodo', 'estudiante__curso__nombre').annotate(
                total=Count('id')
            ).order_by('-total')[:10])

            results['yearly_courses'] = list(qs.annotate(
                periodo=TruncYear('fecha')
            ).values('periodo', 'estudiante__curso__nombre').annotate(
                total=Count('id')
            ).order_by('-total')[:10])

        return results

    rankings = compute_periodic_rankings(atrasos, 'estudiante')

    def fmt_periodo(p):
        try:
            return p.strftime('%Y-%m-%d')
        except Exception:
            return str(p)

    story.append(Paragraph("Estadísticas por Periodo", heading_style))
    story.append(Spacer(1, 8))

    # Tablas: Estudiantes por Semana/Mes
    def add_students_table(title, rows):
        story.append(Paragraph(title, styles['Heading3']))
        if rows:
            data = [["Periodo", "Estudiante", "RUT", "Curso",
                     "Total", "Justificados", "No Justificados"]]
            for r in rows:
                data.append([
                    fmt_periodo(r.get('periodo') or r.get('period')),
                    r.get('estudiante__nombre'),
                    r.get('estudiante__rut'),
                    r.get('estudiante__curso__nombre'),
                    str(r.get('total', 0)),
                    str(r.get('justificados', 0)),
                    str(r.get('no_justificados', 0)),
                ])
            tbl = Table(data, colWidths=[
                        1.1*inch, 1.7*inch, 1.2*inch, 1.5*inch, 0.7*inch, 1.0*inch, 1.1*inch])
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#f9f9f9')]),
            ]))
            story.append(tbl)
        else:
            story.append(
                Paragraph("No hay datos disponibles", styles['Normal']))
        story.append(Spacer(1, 10))

    def add_courses_table(title, rows):
        story.append(Paragraph(title, styles['Heading3']))
        if rows:
            data = [["Periodo", "Curso", "Total"]]
            for r in rows:
                data.append([
                    fmt_periodo(r.get('periodo') or r.get('period')),
                    r.get('estudiante__curso__nombre') or r.get(
                        'curso__nombre') or 'Curso desconocido',
                    str(r.get('total', 0))
                ])
            tbl = Table(data, colWidths=[1.3*inch, 3.0*inch, 1.0*inch])
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#f9f9f9')]),
            ]))
            story.append(tbl)
        else:
            story.append(
                Paragraph("No hay datos disponibles", styles['Normal']))
        story.append(Spacer(1, 10))

    add_students_table("Top Estudiantes por Semana",
                       rankings.get('weekly_students', []))
    story.append(Spacer(1, 10))
    add_students_table("Top Estudiantes por Mes",
                       rankings.get('monthly_students', []))

    # Top estudiantes atrasos
    story.append(
        Paragraph("Top 10 - Estudiantes con Más Atrasos", heading_style))
    if top_estudiantes_atrasos:
        top_atrasos_data = [['Estudiante', 'RUT', 'Curso',
                             'Total Atrasos', 'Justificados', 'No Justificados']]
        for item in top_estudiantes_atrasos:
            top_atrasos_data.append([
                item['estudiante__nombre'],
                item['estudiante__rut'],
                item['estudiante__curso__nombre'],
                str(item['total_atrasos']),
                str(item['atrasos_justificados']),
                str(item['atrasos_no_justificados'])
            ])
        top_atrasos_table = Table(top_atrasos_data, colWidths=[
                                  1.5*inch, 1*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
        top_atrasos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        story.append(top_atrasos_table)
    else:
        story.append(Paragraph("No hay datos disponibles", styles['Normal']))
    story.append(Spacer(1, 20))

    # Generar PDF
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    # Crear respuesta HTTP
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_analytics_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'

    return response


@login_required
def exportar_reporte_excel(request):
    """Exportar reporte en formato Excel"""
    # Obtener los mismos datos que el dashboard
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    curso_id = request.GET.get('curso')
    tipo_incidencia = request.GET.get('tipo_incidencia', 'todos')

    # Aplicar filtros base solo para atrasos
    filtros_atrasos = Q()

    if fecha_inicio:
        filtros_atrasos &= Q(fecha__gte=fecha_inicio)

    if fecha_fin:
        filtros_atrasos &= Q(fecha__lte=fecha_fin)

    if curso_id:
        filtros_atrasos &= Q(estudiante__curso_id=curso_id)

    # Obtener datos
    atrasos = Atraso.objects.filter(filtros_atrasos)

    # Métricas principales
    total_atrasos = atrasos.count()
    atrasos_justificados = atrasos.filter(justificado=True).count()
    atrasos_no_justificados = atrasos.filter(justificado=False).count()

    # Análisis por curso
    atrasos_por_curso = atrasos.values('estudiante__curso__nombre').annotate(
        total=Count('id'),
        justificados=Count('id', filter=Q(justificado=True)),
        no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total')

    # Top estudiantes
    top_estudiantes_atrasos = atrasos.values(
        'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre'
    ).annotate(
        total_atrasos=Count('id'),
        atrasos_justificados=Count('id', filter=Q(justificado=True)),
        atrasos_no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total_atrasos')[:10]

    # Crear workbook de Excel
    wb = Workbook()

    # Hoja 1: Resumen
    ws1 = wb.active
    ws1.title = "Resumen"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092",
                              end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Título
    ws1['A1'] = "REPORTE DE ANALYTICS - ATRASOS"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:E1')

    # Solo fecha de generación
    ws1['A3'] = f"Fecha generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"

    # Métricas principales
    ws1['A9'] = "MÉTRICAS PRINCIPALES"
    ws1['A9'].font = Font(bold=True, size=14)

    headers = ['Métrica', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=10, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    metrics_data = [
        ['Total Atrasos', total_atrasos],
        ['Atrasos Justificados', atrasos_justificados],
        ['Atrasos No Justificados', atrasos_no_justificados],
        ['Porcentaje Justificados',
            f"{round((atrasos_justificados / total_atrasos * 100) if total_atrasos > 0 else 0, 1)}%"],
        ['Porcentaje No Justificados',
            f"{round((atrasos_no_justificados / total_atrasos * 100) if total_atrasos > 0 else 0, 1)}%"],
    ]

    for row, (metric, value) in enumerate(metrics_data, 11):
        ws1.cell(row=row, column=1, value=metric)
        ws1.cell(row=row, column=2, value=value)

    # Hoja 2: Atrasos por Curso
    ws2 = wb.create_sheet("Atrasos por Curso")

    headers = ['Curso', 'Total', 'Justificados', 'No Justificados']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for row, item in enumerate(atrasos_por_curso, 2):
        ws2.cell(row=row, column=1,
                 value=item['estudiante__curso__nombre'] or 'Curso desconocido')
        ws2.cell(row=row, column=2, value=item['total'])
        ws2.cell(row=row, column=3, value=item['justificados'])
        ws2.cell(row=row, column=4, value=item['no_justificados'])

    # Gráfico de barras apiladas: Justificados vs No Justificados por curso
    if ws2.max_row > 2:
        chart_atrasos = BarChart()
        chart_atrasos.type = "col"
        chart_atrasos.grouping = "stacked"
        chart_atrasos.title = "Atrasos por Curso"
        chart_atrasos.y_axis.title = "Cantidad"
        chart_atrasos.x_axis.title = "Curso"
        data_ref = Reference(ws2, min_col=3, max_col=4,
                             min_row=1, max_row=ws2.max_row)
        cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
        chart_atrasos.add_data(data_ref, titles_from_data=True)
        chart_atrasos.set_categories(cats_ref)
        chart_atrasos.height = 12
        chart_atrasos.width = 24
        ws2.add_chart(chart_atrasos, "F2")

    # Rankings por periodo para Excel

    def excel_periodic_rankings(qs):
        res = {}
        if connection.vendor == 'sqlite':
            def week_start(d):
                return d - timedelta(days=d.weekday())

            # Build lists like in PDF
            weekly_students = defaultdict(lambda: defaultdict(int))
            weekly_students_j = defaultdict(lambda: defaultdict(int))
            weekly_students_nj = defaultdict(lambda: defaultdict(int))
            weekly_courses = defaultdict(lambda: defaultdict(int))

            for a in qs:
                w = week_start(a.fecha)
                s_key = (w, a.estudiante.nombre, a.estudiante.rut, getattr(
                    a.estudiante.curso, 'nombre', 'Curso desconocido'))
                c_key = (w, getattr(a.estudiante.curso,
                         'nombre', 'Curso desconocido'))
                weekly_students[s_key]['total'] += 1
                if a.justificado:
                    weekly_students_j[s_key]['j'] += 1
                else:
                    weekly_students_nj[s_key]['nj'] += 1
                weekly_courses[c_key]['total'] += 1

            res['weekly_students'] = sorted([
                {'periodo': k[0], 'estudiante__nombre': k[1], 'estudiante__rut': k[2], 'estudiante__curso__nombre': k[3], 'total': v['total'],
                    'justificados': weekly_students_j.get(k, {}).get('j', 0), 'no_justificados': weekly_students_nj.get(k, {}).get('nj', 0)}
                for k, v in weekly_students.items()
            ], key=lambda x: x['total'], reverse=True)[:10]

            res['weekly_courses'] = sorted([
                {'periodo': k[0], 'estudiante__curso__nombre': k[1],
                    'total': v['total']}
                for k, v in weekly_courses.items()
            ], key=lambda x: x['total'], reverse=True)[:10]

            # Monthly
            monthly_students = defaultdict(lambda: defaultdict(int))
            monthly_students_j = defaultdict(lambda: defaultdict(int))
            monthly_students_nj = defaultdict(lambda: defaultdict(int))
            monthly_courses = defaultdict(lambda: defaultdict(int))

            for a in qs:
                m = a.fecha.replace(day=1)
                s_key = (m, a.estudiante.nombre, a.estudiante.rut, getattr(
                    a.estudiante.curso, 'nombre', 'Curso desconocido'))
                c_key = (m, getattr(a.estudiante.curso,
                         'nombre', 'Curso desconocido'))
                monthly_students[s_key]['total'] += 1
                if a.justificado:
                    monthly_students_j[s_key]['j'] += 1
                else:
                    monthly_students_nj[s_key]['nj'] += 1
                monthly_courses[c_key]['total'] += 1

            res['monthly_students'] = sorted([
                {'periodo': k[0], 'estudiante__nombre': k[1], 'estudiante__rut': k[2], 'estudiante__curso__nombre': k[3], 'total': v['total'],
                    'justificados': monthly_students_j.get(k, {}).get('j', 0), 'no_justificados': monthly_students_nj.get(k, {}).get('nj', 0)}
                for k, v in monthly_students.items()
            ], key=lambda x: x['total'], reverse=True)[:10]

            res['monthly_courses'] = sorted([
                {'periodo': k[0], 'estudiante__curso__nombre': k[1],
                    'total': v['total']}
                for k, v in monthly_courses.items()
            ], key=lambda x: x['total'], reverse=True)[:10]

            # Yearly
            yearly_students = defaultdict(lambda: defaultdict(int))
            yearly_students_j = defaultdict(lambda: defaultdict(int))
            yearly_students_nj = defaultdict(lambda: defaultdict(int))
            yearly_courses = defaultdict(lambda: defaultdict(int))

            for a in qs:
                y = a.fecha.replace(month=1, day=1)
                s_key = (y, a.estudiante.nombre, a.estudiante.rut, getattr(
                    a.estudiante.curso, 'nombre', 'Curso desconocido'))
                c_key = (y, getattr(a.estudiante.curso,
                         'nombre', 'Curso desconocido'))
                yearly_students[s_key]['total'] += 1
                if a.justificado:
                    yearly_students_j[s_key]['j'] += 1
                else:
                    yearly_students_nj[s_key]['nj'] += 1
                yearly_courses[c_key]['total'] += 1

            res['yearly_students'] = sorted([
                {'periodo': k[0], 'estudiante__nombre': k[1], 'estudiante__rut': k[2], 'estudiante__curso__nombre': k[3], 'total': v['total'],
                    'justificados': yearly_students_j.get(k, {}).get('j', 0), 'no_justificados': yearly_students_nj.get(k, {}).get('nj', 0)}
                for k, v in yearly_students.items()
            ], key=lambda x: x['total'], reverse=True)[:10]

            res['yearly_courses'] = sorted([
                {'periodo': k[0], 'estudiante__curso__nombre': k[1],
                    'total': v['total']}
                for k, v in yearly_courses.items()
            ], key=lambda x: x['total'], reverse=True)[:10]

        else:
            res['weekly_students'] = list(qs.annotate(periodo=TruncWeek('fecha')).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'), justificados=Count('id', filter=Q(justificado=True)), no_justificados=Count('id', filter=Q(justificado=False))).order_by('-total')[:10])
            res['monthly_students'] = list(qs.annotate(periodo=TruncMonth('fecha')).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'), justificados=Count('id', filter=Q(justificado=True)), no_justificados=Count('id', filter=Q(justificado=False))).order_by('-total')[:10])
            res['yearly_students'] = list(qs.annotate(periodo=TruncYear('fecha')).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'), justificados=Count('id', filter=Q(justificado=True)), no_justificados=Count('id', filter=Q(justificado=False))).order_by('-total')[:10])
            res['weekly_courses'] = list(qs.annotate(periodo=TruncWeek('fecha')).values(
                'periodo', 'estudiante__curso__nombre').annotate(total=Count('id')).order_by('-total')[:10])
            res['monthly_courses'] = list(qs.annotate(periodo=TruncMonth('fecha')).values(
                'periodo', 'estudiante__curso__nombre').annotate(total=Count('id')).order_by('-total')[:10])
            res['yearly_courses'] = list(qs.annotate(periodo=TruncYear('fecha')).values(
                'periodo', 'estudiante__curso__nombre').annotate(total=Count('id')).order_by('-total')[:10])
        return res

    er = excel_periodic_rankings(atrasos)

    def fmt_p(p):
        try:
            return p.strftime('%Y-%m-%d')
        except Exception:
            return str(p)

    # Crear hojas para rankings por periodo
    def build_students_sheet(title, rows):
        ws = wb.create_sheet(title)
        headers = ['Periodo', 'Estudiante', 'RUT', 'Curso',
                   'Total', 'Justificados', 'No Justificados']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=fmt_p(r.get('periodo')))
            ws.cell(row=row_idx, column=2, value=r.get('estudiante__nombre'))
            ws.cell(row=row_idx, column=3, value=r.get('estudiante__rut'))
            ws.cell(row=row_idx, column=4, value=r.get(
                'estudiante__curso__nombre'))
            ws.cell(row=row_idx, column=5, value=r.get('total'))
            ws.cell(row=row_idx, column=6, value=r.get('justificados'))
            ws.cell(row=row_idx, column=7, value=r.get('no_justificados'))
        return ws

    def build_courses_sheet(title, rows):
        ws = wb.create_sheet(title)
        headers = ['Periodo', 'Curso', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=fmt_p(r.get('periodo')))
            ws.cell(row=row_idx, column=2, value=r.get(
                'estudiante__curso__nombre') or r.get('curso__nombre'))
            ws.cell(row=row_idx, column=3, value=r.get('total'))
        return ws

    ws_week_s = build_students_sheet('Top Est. Semana', er['weekly_students'])
    ws_week_c = build_courses_sheet('Top Cursos Semana', er['weekly_courses'])
    ws_month_s = build_students_sheet('Top Est. Mes', er['monthly_students'])
    ws_month_c = build_courses_sheet('Top Cursos Mes', er['monthly_courses'])
    ws_year_s = build_students_sheet('Top Est. Año', er['yearly_students'])
    ws_year_c = build_courses_sheet('Top Cursos Año', er['yearly_courses'])

    # Hoja 4: Top Estudiantes Atrasos
    ws4 = wb.create_sheet("Top Atrasos")

    headers = ['Estudiante', 'RUT', 'Curso',
               'Total Atrasos', 'Justificados', 'No Justificados']
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for row, item in enumerate(top_estudiantes_atrasos, 2):
        ws4.cell(row=row, column=1, value=item['estudiante__nombre'])
        ws4.cell(row=row, column=2, value=item['estudiante__rut'])
        ws4.cell(row=row, column=3, value=item['estudiante__curso__nombre'])
        ws4.cell(row=row, column=4, value=item['total_atrasos'])
        ws4.cell(row=row, column=5, value=item['atrasos_justificados'])
        ws4.cell(row=row, column=6, value=item['atrasos_no_justificados'])

        # Ajustar ancho de columnas de forma segura

    def adjust_column_widths(worksheet):
        """Ajusta el ancho de las columnas de forma segura"""
        # Convertir número de columna a letra de forma manual
        def get_column_letter(col_num):
            """Convierte número de columna a letra (A, B, C, ..., Z, AA, AB, etc.)"""
            result = ""
            while col_num > 0:
                col_num, remainder = divmod(col_num - 1, 26)
                result = chr(65 + remainder) + result
            return result

        for col_num in range(1, worksheet.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            # Calcular el ancho máximo de la columna
            for row_num in range(1, worksheet.max_row + 1):
                try:
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Ajustar ancho con un mínimo y máximo
            adjusted_width = min(max(max_length + 2, 8), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Aplicar ajuste de columnas a todas las hojas
    for ws in [ws1, ws2, ws4]:
        adjust_column_widths(ws)

    # Guardar archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_analytics_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'

    wb.save(response)
    return response


@login_required
def exportar_reporte_csv(request):
    """Exportar reporte en formato CSV"""
    # Obtener los mismos datos que el dashboard
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    curso_id = request.GET.get('curso')
    tipo_incidencia = request.GET.get('tipo_incidencia', 'todos')

    # Aplicar filtros base solo para atrasos
    filtros_atrasos = Q()

    if fecha_inicio:
        filtros_atrasos &= Q(fecha__gte=fecha_inicio)

    if fecha_fin:
        filtros_atrasos &= Q(fecha__lte=fecha_fin)

    if curso_id:
        filtros_atrasos &= Q(estudiante__curso_id=curso_id)

    # Obtener datos
    atrasos = Atraso.objects.filter(filtros_atrasos)

    # Análisis por curso
    atrasos_por_curso = atrasos.values('estudiante__curso__nombre').annotate(
        total=Count('id'),
        justificados=Count('id', filter=Q(justificado=True)),
        no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total')

    # Top estudiantes
    top_estudiantes_atrasos = atrasos.values(
        'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre'
    ).annotate(
        total_atrasos=Count('id'),
        atrasos_justificados=Count('id', filter=Q(justificado=True)),
        atrasos_no_justificados=Count('id', filter=Q(justificado=False))
    ).order_by('-total_atrasos')[:10]

    # Crear respuesta CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reporte_analytics_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'

    # Configurar writer con encoding UTF-8
    response.write('\ufeff')  # BOM para Excel
    writer = csv.writer(response)

    # Información del reporte
    writer.writerow(['REPORTE DE ANALYTICS - ATRASOS'])
    writer.writerow([])
    writer.writerow(
        [f'Fecha generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
    writer.writerow([])

    # Atrasos por curso
    writer.writerow(['ATRASOS POR CURSO'])
    writer.writerow(['Curso', 'Total', 'Justificados', 'No Justificados'])
    for item in atrasos_por_curso:
        writer.writerow([
            item['estudiante__curso__nombre'] or 'Curso desconocido',
            item['total'],
            item['justificados'],
            item['no_justificados']
        ])
    writer.writerow([])

    # Top estudiantes atrasos
    writer.writerow(['TOP 10 ESTUDIANTES CON MÁS ATRASOS'])
    writer.writerow(['Estudiante', 'RUT', 'Curso',
                    'Total Atrasos', 'Justificados', 'No Justificados'])
    for item in top_estudiantes_atrasos:
        writer.writerow([
            item['estudiante__nombre'],
            item['estudiante__rut'],
            item['estudiante__curso__nombre'],
            item['total_atrasos'],
            item['atrasos_justificados'],
            item['atrasos_no_justificados']
        ])
    writer.writerow([])

    # Rankings por periodo en CSV

    def csv_periodic_rankings(qs):
        res = {}
        if connection.vendor == 'sqlite':
            def week_start(d):
                return d - timedelta(days=d.weekday())
            # Reutilizar lógica simplificada

            def build_counts(key_func_student, key_func_course):
                stu = defaultdict(int)
                stu_j = defaultdict(int)
                stu_nj = defaultdict(int)
                cou = defaultdict(int)
                for a in qs:
                    ks = key_func_student(a)
                    kc = key_func_course(a)
                    stu[ks] += 1
                    if a.justificado:
                        stu_j[ks] += 1
                    else:
                        stu_nj[ks] += 1
                    cou[kc] += 1
                students = [{'periodo': ks[0], 'estudiante__nombre': ks[1], 'estudiante__rut': ks[2], 'estudiante__curso__nombre': ks[3],
                             'total': v, 'justificados': stu_j.get(ks, 0), 'no_justificados': stu_nj.get(ks, 0)} for ks, v in stu.items()]
                courses = [{'periodo': kc[0], 'estudiante__curso__nombre': kc[1],
                            'total': v} for kc, v in cou.items()]
                students.sort(key=lambda x: x['total'], reverse=True)
                courses.sort(key=lambda x: x['total'], reverse=True)
                return students[:10], courses[:10]

            weekly_s, weekly_c = build_counts(lambda a: (week_start(a.fecha), a.estudiante.nombre, a.estudiante.rut, getattr(
                a.estudiante.curso, 'nombre', 'Curso desconocido')), lambda a: (week_start(a.fecha), getattr(a.estudiante.curso, 'nombre', 'Curso desconocido')))
            monthly_s, monthly_c = build_counts(lambda a: (a.fecha.replace(day=1), a.estudiante.nombre, a.estudiante.rut, getattr(
                a.estudiante.curso, 'nombre', 'Curso desconocido')), lambda a: (a.fecha.replace(day=1), getattr(a.estudiante.curso, 'nombre', 'Curso desconocido')))
            yearly_s, yearly_c = build_counts(lambda a: (a.fecha.replace(month=1, day=1), a.estudiante.nombre, a.estudiante.rut, getattr(
                a.estudiante.curso, 'nombre', 'Curso desconocido')), lambda a: (a.fecha.replace(month=1, day=1), getattr(a.estudiante.curso, 'nombre', 'Curso desconocido')))
            res = {
                'weekly_students': weekly_s,
                'weekly_courses': weekly_c,
                'monthly_students': monthly_s,
                'monthly_courses': monthly_c,
                'yearly_students': yearly_s,
                'yearly_courses': yearly_c,
            }
        else:
            res['weekly_students'] = list(qs.annotate(periodo=TruncWeek('fecha')).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'), justificados=Count('id', filter=Q(justificado=True)), no_justificados=Count('id', filter=Q(justificado=False))).order_by('-total')[:10])
            res['monthly_students'] = list(qs.annotate(periodo=TruncMonth('fecha')).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'), justificados=Count('id', filter=Q(justificado=True)), no_justificados=Count('id', filter=Q(justificado=False))).order_by('-total')[:10])
            res['yearly_students'] = list(qs.annotate(periodo=TruncYear('fecha')).values('periodo', 'estudiante__nombre', 'estudiante__rut', 'estudiante__curso__nombre').annotate(
                total=Count('id'), justificados=Count('id', filter=Q(justificado=True)), no_justificados=Count('id', filter=Q(justificado=False))).order_by('-total')[:10])
            res['weekly_courses'] = list(qs.annotate(periodo=TruncWeek('fecha')).values(
                'periodo', 'estudiante__curso__nombre').annotate(total=Count('id')).order_by('-total')[:10])
            res['monthly_courses'] = list(qs.annotate(periodo=TruncMonth('fecha')).values(
                'periodo', 'estudiante__curso__nombre').annotate(total=Count('id')).order_by('-total')[:10])
            res['yearly_courses'] = list(qs.annotate(periodo=TruncYear('fecha')).values(
                'periodo', 'estudiante__curso__nombre').annotate(total=Count('id')).order_by('-total')[:10])
        return res

    cr = csv_periodic_rankings(atrasos)

    def fmt_csv_periodo(p):
        try:
            return p.strftime('%Y-%m-%d')
        except Exception:
            return str(p)

    writer.writerow([])
    writer.writerow(['TOP ESTUDIANTES POR SEMANA'])
    writer.writerow(['Periodo', 'Estudiante', 'RUT', 'Curso',
                    'Total', 'Justificados', 'No Justificados'])
    for r in cr['weekly_students']:
        writer.writerow([fmt_csv_periodo(r['periodo']), r['estudiante__nombre'], r['estudiante__rut'],
                        r['estudiante__curso__nombre'], r['total'], r.get('justificados', 0), r.get('no_justificados', 0)])

    writer.writerow([])
    writer.writerow(['TOP ESTUDIANTES POR MES'])
    writer.writerow(['Periodo', 'Estudiante', 'RUT', 'Curso',
                    'Total', 'Justificados', 'No Justificados'])
    for r in cr['monthly_students']:
        writer.writerow([fmt_csv_periodo(r['periodo']), r['estudiante__nombre'], r['estudiante__rut'],
                        r['estudiante__curso__nombre'], r['total'], r.get('justificados', 0), r.get('no_justificados', 0)])

    # Se excluyen rankings de cursos y anuales del CSV

    return response
